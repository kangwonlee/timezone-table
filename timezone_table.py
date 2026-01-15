#!/usr/bin/env python3
# timezone_table.py
from __future__ import annotations

import argparse
import datetime
import json
import pathlib
import sys

from typing import List, Union


import openpyxl
from openpyxl.styles import Font, PatternFill
from zoneinfo import ZoneInfo, available_timezones, ZoneInfoNotFoundError


# List of cities you want to show
CITY_ZONES = [
    ("New York", "America/New_York"),
    ("London", "Europe/London"),
    ("Singapore", "Asia/Singapore"),
    ("Seoul", "Asia/Seoul"),
    ("Sydney", "Australia/Sydney"),
]

def get_utc_offset(tz: ZoneInfo, dt: datetime.datetime) -> datetime.timedelta:
    """Get UTC offset for sorting."""
    return dt.astimezone(tz).utcoffset() or datetime.timedelta(0)


def format_meeting(
    start: datetime.datetime,
    end: datetime.datetime,
    city: str,
    tz_str: str,
    city_width: int,
) -> str:
    """Format a table row for the meeting in local time."""
    tz = ZoneInfo(tz_str)
    try:
        local_start = start.astimezone(tz)
        local_end = end.astimezone(tz)
    except datetime.zoneinfo.AmbiguousTimeError:
        raise ValueError(f"Ambiguous time in {tz_str} due to DST transition.")
    zone_abbr = local_start.strftime("%Z")
    return f"| {city.ljust(city_width)} | {local_start.strftime('%H:%M')} – {local_end.strftime('%H:%M')} | {zone_abbr} |"


def read_city_zones(cities_file: Union[str, pathlib.Path] = "cities.json") -> List[tuple[str, str]]:
    path = pathlib.Path(cities_file)
    if not path.is_file():
        return CITY_ZONES
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return [(item["city"], item["timezone"]) for item in data]


def parse_datetime(value:str) -> datetime.datetime:
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d %H:%M")  # Customize format as needed
    except ValueError:
        raise argparse.ArgumentTypeError(f"Invalid datetime format: {value}")


def create_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate a Markdown timezone table for a meeting.",
        epilog="Example: python timezone_table.py 2026 1 14 10 0 America/Los_Angeles 60"
    )
    parser.add_argument("year", type=int, help="Year (e.g., 2026)")
    parser.add_argument("month", type=int, help="Month (1-12)")
    parser.add_argument("day", type=int, help="Day (1-31)")
    parser.add_argument("hour", type=int, help="Hour (0-23)")
    parser.add_argument("minute", type=int, help="Minute (0-59)")
    parser.add_argument("timezone", type=str, help="IANA timezone (e.g., America/Los_Angeles)")
    parser.add_argument("duration_minutes", type=int, help="Duration in minutes (positive integer)")
    parser.add_argument("--sort-by-offset", action="store_true", help="Sort cities by UTC offset (west to east)")
    parser.add_argument("--generate-24hour-xlsx", action="store_true", help="Generate 24-hour XLSX table (default: false)")
    parser.add_argument("--output-file", type=str, default="24hour_timezones.xlsx", help="Output file for XLSX")
    return parser


def main(argv:List[str]) -> None:
    parser = create_parser()

    args = parser.parse_args(argv[1:])

    if args.duration_minutes <= 0:
        parser.error("Duration must be positive.")

    try:
        tz = ZoneInfo(args.timezone)
    except ZoneInfoNotFoundError as e:
        print(f"Invalid timezone: {args.timezone}")
        print(e)
        sys.exit(1)

    try:
        meeting_start = datetime.datetime(
            args.year, args.month, args.day, args.hour, args.minute, tzinfo=tz
        )
        meeting_end = meeting_start + datetime.timedelta(minutes=args.duration_minutes)
    except ValueError as e:
        print(f"Invalid date/time: {e}")
        sys.exit(1)

    CITY_ZONES = read_city_zones()

    # Optional: Sort by UTC offset
    if args.sort_by_offset:
        CITY_ZONES.sort(key=lambda x: get_utc_offset(ZoneInfo(x[1]), meeting_start))

    # Dynamic city width
    city_width = max(len(city) for city, _ in CITY_ZONES) + 2  # Padding

    print("# Meeting Time Converter\n")
    print(f"**Original time:** {meeting_start.strftime('%Y-%m-%d %H:%M %Z')} ({args.timezone})")
    print(f"**Duration:** {args.duration_minutes} minutes\n")
    print(f"| {'City'.ljust(city_width)} | Local Time          | Time Zone |")
    print(f"|{'-' * (city_width + 2)}|---------------------|-----------|")

    not_available = []
    for city, tz_str in CITY_ZONES:
        if tz_str not in available_timezones():
            not_available.append((city, tz_str))
            continue
        try:
            print(format_meeting(meeting_start, meeting_end, city, tz_str, city_width))
        except ValueError as e:
            print(f"Error for {city} ({tz_str}): {e}")

    if not_available:
        print("\n**Unavailable timezones:**")
        for city, tz_str in not_available:
            print(f"- {city}: {tz_str}")

    if args.generate_24hour_xlsx:
        base_start = meeting_start.replace(hour=0, minute=0, second=0, microsecond=0)
        write_xl_table(args.timezone, base_start, meeting_start, CITY_ZONES, args.output_file)


def write_xl_table(
    timezone: str,
    base_start: datetime.datetime,
    meeting_start: datetime.datetime,
    city_zones: List[tuple[str, str]],
    output_file: Union[str, pathlib.Path] = "24hour_timezones.xlsx"
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "24-Hour Timezones"

    # Header
    header = [f"Input Hour ({timezone})"] + [f"{city} ({tz_str})" for city, tz_str in city_zones]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Date row
    date_row = ["Date", base_start.strftime('%Y-%m-%d')] + [""] * (len(city_zones) - 1)
    ws.append(date_row)

    # Colors
    green_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")  # Work hours
    gray_fill = PatternFill(start_color="FFA9A9A9", end_color="FFA9A9A9", fill_type="solid")  # Sleep hours

    # Data rows
    for hour in range(24):
        hour_start = base_start + datetime.timedelta(hours=hour)
        row = [f"{hour:02}:00 {hour_start.strftime('%Z')}"]
        for city, tz_str in city_zones:
            if tz_str not in available_timezones():
                row.append("Unavailable")
                continue
            try:
                local_start = hour_start.astimezone(ZoneInfo(tz_str))
                row.append(local_start.strftime('%H:%M %Z'))
            except ValueError as e:
                row.append(f"Error: {e}")
        ws.append(row)

            # Apply colors based on local hour
        for col in range(2, len(row) + 1):
            local_time_str = row[col-1]
            if "Unavailable" in local_time_str or "Error" in local_time_str:
                continue
            local_hour = int(local_time_str.split(':')[0])
            cell = ws.cell(row=ws.max_row, column=col)
            if 9 <= local_hour < 17:  # Work (customize as needed)
                cell.fill = green_fill
            elif 0 <= local_hour < 7 or 22 <= local_hour < 24:  # Sleep
                cell.fill = gray_fill

    wb.save(output_file)
    print(f"Generated 24-hour XLSX: {str(output_file)}")


if __name__ == "__main__":
    main(sys.argv)
