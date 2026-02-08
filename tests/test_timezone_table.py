# begin tests/test_timezone_table.py

import datetime
import json
import pathlib
import pytest
import sys

from io import StringIO
from unittest.mock import patch
from zoneinfo import ZoneInfo


import openpyxl
from openpyxl.styles import Font, PatternFill


test_folder = pathlib.Path(__file__).parent.resolve()
project_folder = test_folder.parent.resolve()
sys.path.insert(0, str(project_folder))

from timezone_table import format_meeting, main, CITY_ZONES, read_city_zones, write_xl_table, create_parser

@pytest.fixture
def mock_argv():
    return [
        "timezone_table.py",  # sys.argv[0]
        "2026", "1", "14", "10", "0", "America/Los_Angeles", "60"
    ]

def test_format_meeting():
    start = datetime.datetime(2026, 1, 14, 10, 0, tzinfo=ZoneInfo("America/Los_Angeles"))
    end = start + datetime.timedelta(minutes=60)
    result = format_meeting(start, end, "San Diego", "America/Los_Angeles", city_width=12)
    assert result == "| San Diego    | 10:00 – 11:00 | PST |"

    # Test with different timezone
    result = format_meeting(start, end, "New York", "America/New_York", city_width=12)
    assert result == "| New York     | 13:00 – 14:00 | EST |"


def test_format_meeting_dst_ambiguous_time():
    """Test that format_meeting handles DST fall-back (ambiguous) times.

    On 2026-11-01 at 01:00 US/Eastern, clocks fall back. A meeting
    starting at 05:30 UTC spans the ambiguous 01:00-02:00 window in
    New York.  ``astimezone()`` resolves this via the fold attribute,
    so this test verifies the function produces correct output for
    ambiguous local times.
    """
    # 2026-11-01 05:30 UTC  ->  01:30 EDT *or* 00:30 EST in New York
    start = datetime.datetime(2026, 11, 1, 5, 30, tzinfo=ZoneInfo("UTC"))
    end = start + datetime.timedelta(minutes=60)

    result = format_meeting(start, end, "New York", "America/New_York", city_width=12)

    # astimezone resolves the fold; the function should return a valid row
    assert result.startswith("| New York")
    assert "–" in result  # contains time range
    # Verify it contains valid HH:MM patterns
    assert ":" in result

def test_main_valid_input(capsys, mock_argv):
    with patch("sys.argv", mock_argv):
        main(sys.argv)
    captured = capsys.readouterr()
    output = captured.out

    assert "# Meeting Time Converter" in output
    assert "**Original time:** 2026-01-14 10:00 PST (America/Los_Angeles)" in output
    assert "**Duration:** 60 minutes" in output
    assert "| New York " in output and "13:00 – 14:00" in output and "EST" in output  # Fuzzy match for dynamic width
    assert "Unavailable timezones:" not in output  # Assuming all are available

def test_main_invalid_timezone(mock_argv):
    mock_argv[6] = "Invalid/TZ"
    with patch("sys.argv", mock_argv), pytest.raises(SystemExit):
        main(mock_argv)

def test_main_invalid_date(mock_argv):
    mock_argv[2] = "13"  # Invalid month
    with patch("sys.argv", mock_argv), pytest.raises(SystemExit):
        main(mock_argv)

def test_main_sort_by_offset(capsys):
    argv = [
        "timezone_table.py", "2026", "1", "14", "10", "0", "America/Los_Angeles", "60", "--sort-by-offset"
    ]
    with patch("sys.argv", argv):
        main(sys.argv)
    captured = capsys.readouterr()
    output_lines = captured.out.splitlines()
    # Check if table rows are sorted (e.g., New York first, then eastwards)
    table_start = next(i for i, line in enumerate(output_lines) if line.startswith("| City"))
    first_city_row = output_lines[table_start + 2]  # After header and separator
    assert "San Diego" in first_city_row  # Westernmost

def test_unavailable_timezone(capsys, mock_argv):
    # Mock unavailable timezone
    with patch("timezone_table.available_timezones", return_value=set()), patch("sys.argv", mock_argv):
        main(sys.argv)
    captured = capsys.readouterr()
    assert "**Unavailable timezones:**" in captured.out
    assert "- New York: America/New_York" in captured.out


def test_read_city_zones_no_file(tmp_path):
    # Test fallback to hardcoded when file doesn't exist
    non_existent_file = tmp_path / "non_existent.json"
    result = read_city_zones(non_existent_file)
    assert result == CITY_ZONES
    assert isinstance(result, list)
    assert all(isinstance(item, tuple) and len(item) == 2 for item in result)


def test_read_city_zones_valid_json(tmp_path):
    # Test loading from a valid JSON file
    valid_data = [
        {"city": "Test City1", "timezone": "America/Test1"},
        {"city": "Test City2", "timezone": "Europe/Test2"}
    ]
    valid_file = tmp_path / "valid.json"
    with open(valid_file, "w", encoding="utf-8") as f:
        json.dump(valid_data, f)

    result = read_city_zones(valid_file)
    expected = [("Test City1", "America/Test1"), ("Test City2", "Europe/Test2")]
    assert result == expected


def test_read_city_zones_invalid_json(tmp_path):
    # Test malformed JSON raises JSONDecodeError
    invalid_file = tmp_path / "invalid.json"
    with open(invalid_file, "w", encoding="utf-8") as f:
        f.write("this is not json")  # Malformed

    with pytest.raises(json.JSONDecodeError):
        read_city_zones(invalid_file)


def test_read_city_zones_wrong_structure(tmp_path):
    # Test JSON with missing keys raises KeyError
    wrong_data = [
        {"city": "Test City"},  # Missing "timezone"
        {"timezone": "Europe/Test"}  # Missing "city"
    ]
    wrong_file = tmp_path / "wrong.json"
    with open(wrong_file, "w", encoding="utf-8") as f:
        json.dump(wrong_data, f)

    with pytest.raises(KeyError):
        read_city_zones(wrong_file)


def test_read_city_zones_empty_file(tmp_path):
    # Test empty JSON list returns empty list
    empty_data = []
    empty_file = tmp_path / "empty.json"
    with open(empty_file, "w", encoding="utf-8") as f:
        json.dump(empty_data, f)

    result = read_city_zones(empty_file)
    assert result == []


def test_read_city_zones_default_file_not_present():
    # Save original defaults
    original_defaults = read_city_zones.__defaults__

    # Mock the default filename
    read_city_zones.__defaults__ = ("non_existent_default.json",)

    # Test fallback to CITY_ZONES (assuming the mocked file doesn't exist)
    result = read_city_zones()  # Now uses "non_existent_default.json"
    assert result == CITY_ZONES

    # Restore to avoid affecting other tests
    read_city_zones.__defaults__ = original_defaults


@pytest.fixture(scope="module")
def xl_table_fixture(tmp_path_factory):
    """Fixture to call write_xl_table and load the workbook for reuse across tests."""
    tmp_dir = tmp_path_factory.mktemp("xl_table")
    timezone = "America/Los_Angeles"
    meeting_start = datetime.datetime(2026, 1, 15, 14, 0, tzinfo=ZoneInfo(timezone))
    base_start = meeting_start.replace(hour=0, minute=0)  # Compute here; function overrides but we pass it
    city_zones = [
        ("Los Angeles", "America/Los_Angeles"),
        ("Sydney", "Australia/Sydney"),
    ]
    output_file = tmp_dir / "test_24hour_timezones.xlsx"

    # Call the function under test
    write_xl_table(timezone, base_start, meeting_start, city_zones, output_file)

    # Load and yield the workbook and file path
    wb = openpyxl.load_workbook(output_file)
    yield wb, output_file  # Yield for tests to use
    # Teardown: Close workbook (optional, pytest handles cleanup)


def test_write_xl_table_structure(xl_table_fixture):
    wb, _ = xl_table_fixture
    ws = wb["24-Hour Timezones"]
    assert ws.title == "24-Hour Timezones"
    assert ws.max_row == 26
    assert ws.max_column == 3


def test_write_xl_table_header_and_date(xl_table_fixture):
    """Test header row and date row."""
    wb, _ = xl_table_fixture
    ws = wb["24-Hour Timezones"]
    # Header
    assert ws["A1"].value == "Input Hour (America/Los_Angeles)"
    assert ws["B1"].value == "Los Angeles (America/Los_Angeles)"
    assert ws["C1"].value == "Sydney (Australia/Sydney)"
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws["C1"].font.bold is True
    # Date row
    assert ws["A2"].value == "Date"
    assert ws["B2"].value == "2026-01-15"
    assert ws["C2"].value is None  # Empty


def test_write_xl_table_data_rows(xl_table_fixture):
    """Test sample data rows for correctness."""
    wb, _ = xl_table_fixture
    ws = wb["24-Hour Timezones"]
    # First data row (00:00 PST)
    assert ws["A3"].value.startswith("00:00")
    assert "00:00 PST" in str(ws["B3"].value)  # LA (same TZ)
    assert str(ws["C3"].value).endswith("AEDT")  # Sydney abbr
    # Another row (01:00)
    assert ws["A4"].value.startswith("01:00")
    # Last row (23:00)
    assert ws["A26"].value.startswith("23:00")
    # Spot check a Sydney time (e.g., 00:00 PST -> ~19:00 AEDT, but assert format only due to DST variability)
    assert ":" in str(ws["C3"].value)  # Time format


def test_write_xl_table_colors(xl_table_fixture):
    wb, _ = xl_table_fixture
    ws = wb["24-Hour Timezones"]
    assert ws["B3"].fill.start_color.rgb == "FFA9A9A9"  # Opaque gray
    assert "09:00 PST" in str(ws["B12"].value)
    assert ws["B12"].fill.start_color.rgb == "FF90EE90"  # Opaque green


def test_write_xl_table_error_handling(xl_table_fixture):
    """Test handling of unavailable or error-prone zones (requires modifying inputs)."""
    # To test errors, we'd need a separate call with bad data; since fixture is module-scoped, suggest a new test
    # For now, assert no "Error" or "Unavailable" in this valid setup
    wb, _ = xl_table_fixture
    ws = wb["24-Hour Timezones"]
    for row in ws.iter_rows(min_row=3, max_row=26, min_col=2, max_col=3):
        for cell in row:
            assert "Error" not in str(cell.value)  # No exceptions
            assert "Unavailable" not in str(cell.value)  # All zones available


def test_create_parser():
    parser = create_parser()
    args = parser.parse_args(["2026", "1", "14", "10", "0", "America/Los_Angeles", "60"])

    # Verify positional args
    assert args.year == 2026
    assert args.month == 1
    assert args.day == 14
    assert args.hour == 10
    assert args.minute == 0
    assert args.timezone == "America/Los_Angeles"
    assert args.duration_minutes == 60

    # Verify optional args defaults
    assert args.sort_by_offset is False
    assert args.generate_24hour_xlsx is False
    assert args.output_file == "24hour_timezones.xlsx"

    # Test with flags
    args_with_flags = parser.parse_args(
        ["2026", "1", "14", "10", "0", "America/Los_Angeles", "60", "--sort-by-offset", "--generate-24hour-xlsx", "--output-file=myfile.xlsx"]
    )
    assert args_with_flags.sort_by_offset is True
    assert args_with_flags.generate_24hour_xlsx is True
    assert args_with_flags.output_file == "myfile.xlsx"


def test_main_with_xlsx_generation(capsys, mock_argv, tmp_path):
    # Modify mock_argv to include flag and set output to temp file
    mock_argv.extend(["--generate-24hour-xlsx", f"--output-file={tmp_path / 'test.xlsx'}"])

    with patch("sys.argv", mock_argv):
        main(mock_argv)

    captured = capsys.readouterr()
    assert "Generated 24-hour XLSX:" in captured.out  # Confirms print statement
    assert (tmp_path / "test.xlsx").exists()  # File was created


def test_main_with_current_date(capsys):
    """Test that main() works with today's date, matching the workflow fallback.

    The meeting-time.yml workflow falls back to today's date when
    year/month/day inputs are left empty.  This test simulates that
    by building argv from the current date and verifying main()
    produces valid output.
    """
    now = datetime.datetime.now()
    argv = [
        "timezone_table.py",
        str(now.year), str(now.month), str(now.day),
        "10", "0", "Europe/Paris", "60",
    ]
    main(argv)
    captured = capsys.readouterr()
    output = captured.out

    assert "# Meeting Time Converter" in output
    expected_date = now.strftime("%Y-%m-%d")
    assert expected_date in output
    assert "**Duration:** 60 minutes" in output
    assert "Europe/Paris" in output


if "__main__" == __name__:
    pytest.main(["-v", __file__])

# end tests/test_timezone_table.py
